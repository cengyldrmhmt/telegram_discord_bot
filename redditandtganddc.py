import time
import discord
import asyncio
import config
import datetime
import openpyxl
from xlrd import open_workbook
#telegram
import configparser
from telethon import TelegramClient
from telethon.errors import SessionPasswordNeededError
from telethon.tl.functions.messages import (GetHistoryRequest)
from telethon.tl.types import (
    PeerChannel
)


#TELEGRAM API
config_tg = configparser.ConfigParser()
config_tg.read("config.ini")

api_id = config_tg['Telegram']['api_id']
api_hash = config_tg['Telegram']['api_hash']

api_hash = str(api_hash)

phone = config_tg['Telegram']['phone']
username_tg = config_tg['Telegram']['username']

# Create the client and connect
client1 = TelegramClient(username_tg, api_id, api_hash)


#Discord Bot Ana ayarları
DCTOKEN ='{Discord token}' # Discord tokenı buraya yapıştıracaksınız
client = discord.Client(intents=discord.Intents.default());



Postlinks = []


async def send_TG_message(phone,message,image):
    await client1.start()
    print("Client Created")
    # Ensure you're authorized
    if await client1.is_user_authorized() == False:
        await client1.send_code_request(phone)
        try:
            await client1.sign_in(phone, input('Enter the code: '))  # Telegram uygulamanıza kod gelecek 
        except SessionPasswordNeededError:
            await client1.sign_in(password=input('Password: '))

    me = await client1.get_me()
    await client1.send_file('https://t.me/{kullanıcıadı}',image,caption=message) # ÖRNEK TELEGRAM KANAL LİNKİ : https://t.me/badgirlsmegas
    await asyncio.sleep(2)


@client.event
async def on_ready():
    while True:
        

        path = "TGPOSTLAR.xlsx"  # paylaşılacak içeriklerin çekileceği excel.
 
        wb_obj = openpyxl.load_workbook(path)
         
        sheet_obj = wb_obj.active
        m_row = sheet_obj.max_row
         


        await asyncio.sleep(5) # 5 SN post var mı yok mu kontrol eder.
        print ("\n\nPOST VAR MI YOK MU KONTROL EDİYORUM (5 sn)","\n")
        for i in range(1, m_row + 1):
            ex_title=sheet_obj.cell(row = i, column = 1).value
            ex_image=sheet_obj.cell(row = i, column = 2).value
            ex_megalink=sheet_obj.cell(row = i, column = 3).value
            PostTime=sheet_obj.cell(row = i, column = 4).value
            with open("TGPaylasilanlinks.txt", "r") as file3:  #Paylasilansublar.txt elle oluştur.
                Postlinks = file3.read()
                Postlinks = Postlinks.split("\n")
                Postlinks = list(filter(None, Postlinks))
                if ex_megalink not in Postlinks :
                    print("\n\n YENİ İÇERİK BULUNDU -> ",ex_megalink,"\n")

                    with open ("TGPaylasilanlinks.txt", "a") as f:
                        
                        f.write(ex_megalink + "\n")
                        
                        image ='TGimgs/'+ex_image
                        

                        title = ex_title+"\n\n\n 🟥🟥MEGA PACK LINK : "
                        megalink = ex_megalink

                        emojis="\n\n🌕🌖🌗🌘🌑🌒🌓🌔🌕🌖🌗🌘🌑🌒🌓🌔🌕"

                        #bütün (yazı olan) içerikleri burada birleştiriyorsun.
                        TGallText = title+megalink+emojis
                        imageFile = {'photo':open(image,'rb')}

                        #discord için
                        dctitle = "\n\n\n"+ex_title+"\n\n\nMEGA PACK LINK : "
                        dcmegalink = "<"+ex_megalink+">"
                        
                        allText = dctitle+dcmegalink+emojis

                        
                        #Telegram Paylaşımın yapıltıdığı yer
                        async with client1:
                            await send_TG_message(phone,TGallText,image)
                        #telegram kanal   


                        file = discord.File(image, filename = image)
                        channel = client.get_channel(123123123123) # Discord text kanal idsi gelecek.
                        await channel.send(allText + "@everyone" ,file = file)
                        
                      
                      
                        #Paylaşımın yapıltıdığı yer
                        
                        x = datetime.datetime.now()
                        x = int(x.strftime("%H"))
                        with open ("TGPaylasilanlinks.txt", "a") as f:
                             f.write(ex_megalink + "\n")
                             f.close()
                        if(5>=x>=2):  # gece 2 ile 5 arası 3 saatte bir atsın.
                            await asyncio.sleep(10800)   
                        elif(9>=x>=6):   # sabah 6 ile 9 arası 2 saatte bir atsın.
                            await asyncio.sleep(7200) 
                        else:    
                            if (PostTime==None):     
                                 print("süre 3600 sn")
                                 await asyncio.sleep(3600)  # 4 sütuna bir şey yazmadıysanız 1 saatte bir post yollayacak.
                            else:
                                 print("Süre ",PostTime)
                                 await asyncio.sleep(int(PostTime))
                        
                        
                    
                    
             


client.run(DCTOKEN)
#client.logout()

    